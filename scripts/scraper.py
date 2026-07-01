"""Playwright-based scraper for both static and SPA pages."""

from playwright.sync_api import sync_playwright, Page, Browser
from bs4 import BeautifulSoup


class Scraper:
    def __init__(self, headless: bool = True, timeout: int = 30000):
        self.headless = headless
        self.timeout = timeout
        self._playwright = None
        self._browser: Browser | None = None
        self._page_pool: list[Page] = []

    def start(self):
        self._playwright = sync_playwright().start()
        self._browser = self._playwright.chromium.launch(headless=self.headless)

    def stop(self):
        for p in self._page_pool:
            try: p.close()
            except: pass
        self._page_pool.clear()
        if self._browser:
            self._browser.close()
        if self._playwright:
            self._playwright.stop()

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, *args):
        self.stop()

    def borrow_page(self) -> Page:
        """Get a reusable page. For detail scraping to avoid new_page/close per article."""
        if self._page_pool:
            return self._page_pool.pop()
        return self._browser.new_page()

    def return_page(self, page: Page):
        """Return page to pool for reuse."""
        self._page_pool.append(page)

    def navigate(self, page: Page, url: str, wait_ms: int = 3000,
                  wait_for: str | list[str] | None = None) -> str:
        """Navigate existing page to url, return HTML. Much faster than new_page+close.

        Uses "commit" (not "domcontentloaded") because some legacy gov sub-sites
        never fire DOMContentLoaded, causing 30s timeouts.
        If wait_for is provided, waits up to 8s for the selector before falling
        back to wait_ms.
        """
        page.goto(url, wait_until="commit", timeout=self.timeout)
        if wait_for:
            selectors = [wait_for] if isinstance(wait_for, str) else wait_for
            for sel in selectors:
                try:
                    page.wait_for_selector(sel, timeout=8000)
                    break
                except Exception:
                    continue
        page.wait_for_timeout(wait_ms)
        return page.content()

    def fetch_page(self, url: str, wait_for: str | list[str] | None = None,
                   wait_ms: int = 3000, retries: int = 3) -> str:
        """Navigate to url, return HTML content.

        wait_for: CSS selector(s) to wait for. If list, tries each until one matches.
        wait_ms: fallback wait in ms after navigation.
        retries: max retry attempts on transient network errors (ERR_NETWORK_CHANGED etc.)
        """
        last_error = None
        for attempt in range(retries):
            page: Page = self._browser.new_page()
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=self.timeout)
                if wait_for:
                    selectors = [wait_for] if isinstance(wait_for, str) else wait_for
                    found = False
                    for sel in selectors:
                        try:
                            page.wait_for_selector(sel, timeout=8000)
                            found = True
                            break
                        except Exception:
                            continue
                    if not found:
                        page.wait_for_timeout(wait_ms)
                else:
                    page.wait_for_timeout(wait_ms)
                return page.content()
            except Exception as e:
                last_error = e
                if attempt < retries - 1 and ("ERR_NETWORK_CHANGED" in str(e) or "ERR_INTERNET_DISCONNECTED" in str(e) or "ERR_CONNECTION" in str(e)):
                    print(f"  ⚠ 网络瞬时错误，重试 {attempt + 2}/{retries} ...")
                    import time
                    time.sleep(2)
                else:
                    raise
            finally:
                page.close()
        raise last_error

    def click_load_more(self, url: str, button_selector: str,
                        clicks: int = 3, wait_ms: int = 1500) -> str:
        """For 'load more' pagination: click button N times, return final HTML."""
        page: Page = self._browser.new_page()
        try:
            page.goto(url, wait_until="commit", timeout=self.timeout)
            page.wait_for_timeout(wait_ms)
            for _ in range(clicks):
                try:
                    btn = page.query_selector(button_selector)
                    if btn and btn.is_visible():
                        btn.click()
                        page.wait_for_timeout(wait_ms)
                    else:
                        break
                except Exception:
                    break
            return page.content()
        finally:
            page.close()

    def click_next_pages(self, url: str, link_selector: str,
                         pages: int = 3, wait_ms: int = 1500) -> list[str]:
        """For numbered pagination: return list of HTML strings for N pages."""
        results = []
        page: Page = self._browser.new_page()
        try:
            page.goto(url, wait_until="commit", timeout=self.timeout)
            page.wait_for_timeout(wait_ms)
            results.append(page.content())
            for _ in range(pages - 1):
                try:
                    next_link = page.query_selector(link_selector)
                    if next_link and next_link.is_visible():
                        next_link.click()
                        page.wait_for_timeout(wait_ms)
                        results.append(page.content())
                    else:
                        break
                except Exception:
                    break
            return results
        finally:
            page.close()

    @staticmethod
    def to_soup(html: str) -> BeautifulSoup:
        return BeautifulSoup(html, "lxml")

    @staticmethod
    def download_attachments(attachment_links: list[tuple[str, str]]) -> tuple[str, list[dict]]:
        """Download attachment files, extract text + HTML tables.

        attachment_links: list of (url, filename) tuples
        Returns (extracted_text, html_tables) where html_tables is
        [{filename, html, sheet_name}, ...] for spreadsheet files.
        """
        import tempfile, os, re
        try:
            import requests as req
        except ImportError:
            print("[Scraper] requests not available, skipping attachments")
            return "", []

        extracted_parts = []
        html_tables = []
        seen_filenames = set()  # dedup identical attachments
        for url, filename in attachment_links:
            try:
                if filename in seen_filenames:
                    continue
                seen_filenames.add(filename)
                print(f"  [Scraper] 下载附件: {filename}")
                resp = req.get(url, timeout=30, headers={
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
                })
                resp.raise_for_status()
                content = resp.content
                ext = os.path.splitext(filename)[1].lower()

                text = ""
                html_table = None
                if ext in (".pdf",):
                    text = Scraper._extract_pdf(content)
                elif ext in (".xlsx", ".xls"):
                    text = Scraper._extract_xlsx(content, ext)
                    html_table = Scraper._extract_spreadsheet_to_html(content, ext, filename)
                elif ext in (".docx",):
                    text = Scraper._extract_docx(content)
                elif ext in (".csv", ".txt", ".md"):
                    text = content.decode("utf-8", errors="replace")[:8000]
                else:
                    try:
                        text = content.decode("utf-8")[:8000]
                    except UnicodeDecodeError:
                        print(f"  [Scraper] 无法解析: {filename}")
                        continue

                if text.strip():
                    extracted_parts.append(f"--- {filename} ---\n{text}")
                if html_table:
                    html_tables.extend(html_table)
            except Exception as e:
                print(f"  [Scraper] 附件下载失败 {filename}: {e}")

        return "\n\n".join(extracted_parts), html_tables

    @staticmethod
    def _extract_pdf(content: bytes) -> str:
        try:
            import pdfplumber, io
            text_parts = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages[:5]:  # first 5 pages
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
            return "\n".join(text_parts)[:8000]
        except ImportError:
            pass
        try:
            from PyPDF2 import PdfReader
            import io
            reader = PdfReader(io.BytesIO(content))
            return "\n".join(p.extract_text() or "" for p in reader.pages[:5])[:8000]
        except ImportError:
            return ""

    @staticmethod
    def _extract_xlsx(content: bytes, ext: str) -> str:
        if ext == ".csv":
            return content.decode("utf-8", errors="replace")[:8000]

        # Old .xls (BIFF/OLE2) — use xlrd
        if ext == ".xls":
            try:
                import xlrd, io
                wb = xlrd.open_workbook(file_contents=content)
                parts = []
                for sn in wb.sheet_names()[:3]:
                    ws = wb.sheet_by_name(sn)
                    parts.append(f"[Sheet: {sn}]")
                    rows = []
                    for r in range(min(ws.nrows, 30)):
                        rows.append("\t".join(
                            str(ws.cell_value(r, c)) if ws.cell_value(r, c) != "" else ""
                            for c in range(ws.ncols)))
                    parts.append("\n".join(rows))
                return "\n".join(parts)[:8000]
            except ImportError:
                return ""
            except Exception as e:
                print(f"  [Scraper] XLS解析失败: {e}")
                return ""

        # .xlsx (Open XML) — use openpyxl
        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            parts = []
            for sn in wb.sheetnames[:3]:
                ws = wb[sn]
                parts.append(f"[Sheet: {sn}]")
                rows = []
                for row in ws.iter_rows(max_row=30, values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
                parts.append("\n".join(rows))
            return "\n".join(parts)[:8000]
        except ImportError:
            return ""
        except Exception as e:
            print(f"  [Scraper] XLSX解析失败: {e}")
            return ""

    @staticmethod
    def _extract_spreadsheet_to_html(content: bytes, ext: str, filename: str) -> list[dict]:
        """Convert spreadsheet to styled HTML tables. Returns [{filename, sheet_name, html}, ...]."""
        tables = []
        try:
            if ext == ".xls":
                import xlrd
                wb = xlrd.open_workbook(file_contents=content)
                for sn in wb.sheet_names():
                    ws = wb.sheet_by_name(sn)
                    if ws.nrows == 0:
                        continue
                    rows_html = []
                    for r in range(ws.nrows):
                        cells = []
                        tag = "th" if r == 0 else "td"
                        for c in range(ws.ncols):
                            v = ws.cell_value(r, c)
                            s = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
                            if s == "0.0":
                                s = "0"
                            cells.append(f"<{tag}>{s or '—'}</{tag}>")
                        rows_html.append(f"<tr>{''.join(cells)}</tr>")
                    label = f"{filename}" if len(wb.sheet_names()) == 1 else f"{filename} · {sn}"
                    tables.append({
                        "filename": filename,
                        "sheet_name": sn,
                        "html": f"<table class='data-table'><caption>{label}</caption>"
                                f"<tbody>{''.join(rows_html)}</tbody></table>",
                    })
            else:  # .xlsx
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    rows = list(ws.iter_rows(max_row=100, values_only=True))
                    if not rows:
                        continue
                    rows_html = []
                    for i, row in enumerate(rows):
                        cells = []
                        tag = "th" if i == 0 else "td"
                        for c in row:
                            v = c if c is not None else ""
                            s = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
                            if s == "0.0":
                                s = "0"
                            cells.append(f"<{tag}>{s or '—'}</{tag}>")
                        rows_html.append(f"<tr>{''.join(cells)}</tr>")
                    label = f"{filename}" if len(wb.sheetnames) == 1 else f"{filename} · {sn}"
                    tables.append({
                        "filename": filename,
                        "sheet_name": sn,
                        "html": f"<table class='data-table'><caption>{label}</caption>"
                                f"<tbody>{''.join(rows_html)}</tbody></table>",
                    })
        except ImportError:
            pass
        except Exception as e:
            print(f"  [Scraper] 表格转换失败 {filename}: {e}")
        return tables

    @staticmethod
    def _extract_docx(content: bytes) -> str:
        # Try python-docx first (handles .docx / Open XML)
        try:
            from docx import Document
            import io
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs[:50])[:8000]
        except ImportError:
            pass
        except Exception:
            pass  # May be old .doc format (OLE2), try fallback

        # Fallback for old .doc format (OLE2/CFB) — use macOS textutil
        if content[:4] == b'\xd0\xcf\x11\xe0':
            return Scraper._extract_ole_doc(content)
        return ""

    @staticmethod
    def _extract_ole_doc(content: bytes) -> str:
        """Extract text from old .doc (OLE2) using macOS textutil or olefile."""
        import subprocess, tempfile, os
        # Method 1: macOS textutil (built-in)
        try:
            with tempfile.NamedTemporaryFile(suffix='.doc', delete=False) as f:
                f.write(content)
                tmp = f.name
            result = subprocess.run(
                ['textutil', '-convert', 'txt', tmp, '-stdout'],
                capture_output=True, text=True, timeout=30,
            )
            os.unlink(tmp)
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout[:8000]
        except Exception:
            pass

        # Method 2: olefile + manual extraction
        try:
            import olefile
            ole = olefile.OleFileIO(content)
            if ole.exists('WordDocument'):
                # Read the WordDocument stream — crude but works for simple docs
                stream = ole.openstream('WordDocument').read()
                # Try to extract readable text (UTF-16LE encoded in OLE docs)
                text = stream.decode('utf-16-le', errors='ignore')
                # Filter to printable CJK + ASCII
                import re
                cleaned = re.sub(r'[^一-鿿　-〿＀-￯a-zA-Z0-9\s.,;:!?()（）【】《》、。，；：！？\d%+\-*/=]', '', text)
                if len(cleaned) > 100:
                    ole.close()
                    return cleaned[:8000]
                ole.close()
        except ImportError:
            pass
        except Exception:
            pass

        return ""
